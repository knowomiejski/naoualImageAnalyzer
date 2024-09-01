

import scyjava as sj
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook
from os import popen
from os import mkdir
from os import listdir
from os import fsdecode
from os import path
import time
from models.directory import Directory
import imagej
import glob
import csv


class ImageAnalyzer:


    def __init__(self):
        self.outPutDirectory = "data/outputs"
        self.ij = imagej.init("imagej/Fiji.app", mode="interactive")
        # self.ij = imagej.init("imagej/Fiji.segmentation.app", mode="interactive")
        # self.ij.IJ.run("Console", "")
        # self.ij.IJ.run("Record...", "");
        print("STARTING IMAGEJ...")
        print(self.ij.getVersion())

    def setOutputDir(self, outputDir):
        self.outPutDirectory = outputDir

    def getOutputDir(self):
        return path.join(self.outPutDirectory)

    def summariesToExcel(self, directory: Directory):
        if directory.hasBeenAnalyzed == False:
            print("NOT ANALYZED: ", directory.getDirectoryPath())
            return
        absolutePath = path.abspath(directory.getDirectoryPath())
        summariesDir = path.join(str(absolutePath), "Best Segmentation", "Combined Files")
        allSummaryFileValuesCSV = path.join(summariesDir, "All Summary File Values.csv")
        if (path.exists(path.join(summariesDir, "All_Summary_File_Values.xlsx")) == False):
            wb = Workbook()
            ws = wb.worksheets[0]
            wb.save(path.join(summariesDir, "All_Summary_File_Values.xlsx"))

        csvfile = open(allSummaryFileValuesCSV, "r")
        reader = csv.reader(csvfile, delimiter=",")
        data = list(reader)

        wb = load_workbook(path.join(summariesDir, "All_Summary_File_Values.xlsx"))
        ws = wb.worksheets[0]

        for x in data:
            rowRaw = x
            rowToAdd = []
            for row in rowRaw:
                if self.is_float(row):
                    rowToAdd.append(float(row))
                else:
                    rowToAdd.append(row)
            ws.append(rowToAdd)

        wb.save(path.join(summariesDir, "All_Summary_File_Values.xlsx"))
        csvfile.close()

    def is_float(self, string):
        if string.replace("-", "").replace(".", "").isnumeric():
            return True
        else:
            return False

    def runDiameterJOnBestSegmentation(self, directory: Directory):
        print("For directory", directory.getHasBeenSegmented())
        print("Strat", directory.getFullStrategy())
        absolutePath = path.abspath(directory.getDirectoryPath())
        segmentedDir = path.join(str(absolutePath), "Segmented Images")
        segmentedDir = path.join(segmentedDir, "")
        bestDir = path.join(str(absolutePath), "Best Segmentation")
        bestDir = path.join(bestDir, "")

        for file in listdir(segmentedDir):
            filename = fsdecode(file)
            if filename.endswith(directory.getFullStrategy() + ".tif"):
                print("copy \"" + bestDir + filename)
                popen("copy \"" + segmentedDir + filename + "\" \""+ bestDir + filename + "\"")


        firstFile = ""
        for file in listdir(absolutePath):
            filename = fsdecode(file)
            if filename.endswith('.tif') != False:
                firstFile = filename
            else:
                continue

        BF = sj.jimport('loci.plugins.BF')
        options = sj.jimport('loci.plugins.in.ImporterOptions')() # import and initialize ImporterOptions
        options.setOpenAllSeries(True)
        options.setVirtual(True)
        options.setId(path.join(absolutePath, firstFile))
        imp = BF.openImagePlus(options)[0]
        # imp.show()
        # imageInfo = imp.getProperties()
        # print("imageInfo", imageInfo)

        self.ij.IJ.run("DiameterJ 1-018", "orientation=OrientationJ do=Yes length=2424 length_0=103.10 do_0=No min.=1 max.=255 do_1=Yes do_2=Yes choose=[" + path.join(bestDir, "").replace("\\", "/") + "]")
        directory.setHasBeenAnalyzed(True)

        for csvfile in glob.glob(path.join(bestDir, "Histograms")  + "*.csv"):
            workbook = Workbook(csvfile[:-4] + '.xlsx')
            worksheet = workbook.add_worksheet()
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.write(r, c, col)
            workbook.close()

        for csvfile in glob.glob(path.join(bestDir + "Summaries") + "*.csv"):
            workbook = Workbook(csvfile[:-4] + '.xlsx')
            worksheet = workbook.add_worksheet()
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.write(r, c, col)
            workbook.close()


    def runSegmentation(self, directory: Directory):
        print("Directory...", directory)
        absolutePath = path.abspath(directory.getDirectoryPath())
        time.sleep(6)
        print("Analyzing directory..." + absolutePath)
        i = 0
        firstFile = ""
        for file in listdir(absolutePath):
            filename = fsdecode(file)
            if filename.endswith('.tif') != False:
                firstFile = filename
                break
            else:
                continue
        print("FILE NAME:" + firstFile)
        print("FILE NAME:" + path.join(absolutePath, firstFile))
        self.ij.ui().showUI()
        BF = sj.jimport('loci.plugins.BF')
        options = sj.jimport('loci.plugins.in.ImporterOptions')() # import and initialize ImporterOptions
        options.setOpenAllSeries(True)
        options.setVirtual(True)
        options.setId(path.join(absolutePath, firstFile))
        imp = BF.openImagePlus(options)[0]
        imp.show()
        width = imp.getWidth()
        heigth = imp.getHeight()
        print("IMP WIDTH=",  imp.getWidth())
        print("IMP HEIGHT=", imp.getHeight())

        self.ij.IJ.run("DiameterJ Segment", "do=No image=" + str(width) +" image_0=" + str(heigth) + " traditional stat. mixed do_0=Yes choose=" + path.join(absolutePath, ""))
        # self.ij.IJ.run("DiameterJ Segment", "do=No image=1216 image_0=1216 traditional stat. mixed do_0=Yes choose=A:\\Development\\NaoualImageAnalyzer\\data\\outputs\\Experiment-1021\\")
        # self.ij.IJ.run("DiameterJ Segment", "do=No image=1216 image_0=1216 traditional stat. mixed do_0=Yes choose=A:/Development/NaoualImageAnalyzer/data/outputs/Experiment-1021/");

        directory.setHasBeenSegmented(True)

    def splitFile(self, fileLocation):
        BF = sj.jimport('loci.plugins.BF')
        options = sj.jimport('loci.plugins.in.ImporterOptions')() # import and initialize ImporterOptions
        options.setId(fileLocation)
        options.setVirtual(False)
        options.setSplitChannels(True)

        options.setCBegin(0,1)
        options.setCEnd(0,1)
        options.setCStep(0,1)

        options.setZBegin(0,3)
        options.setZEnd(0,12)
        options.setZStep(0,1)

        options.setTBegin(0,0)
        options.setTEnd(0,0)
        options.setTStep(0,1)
        imp = BF.openImagePlus(options)[0]
        imp.show();
        self.ij.IJ.run(imp, "8-bit", "");
        self.ij.IJ.run("Stack to Images", "");
        counter = 0
        while(True):
            counter = counter + 1
            activeImage = self.ij.py.active_imageplus()
            if (activeImage == None):
                break
            else:
                print(activeImage.getFileInfo())
                splitFileLocation = str(str(fileLocation).split('/')[-1].split('.')[0]).replace(" ", "")
                if (path.exists(path.join(self.getOutputDir(), splitFileLocation)) == False):
                    mkdir(path.join(self.getOutputDir(), splitFileLocation))
                stringlengthOfCounter = len(str(counter))
                amountOfZeroes= 4 - stringlengthOfCounter
                zeroes = '0'*amountOfZeroes
                self.ij.IJ.saveAs(activeImage, "Tiff", path.join(self.getOutputDir(), splitFileLocation, splitFileLocation + "_" + zeroes + str(counter) + ".tif"))
                activeImage.close()

    def testLoadImage(self):
        image_url = "data/Experiment-1022.czi"
        BF = sj.jimport('loci.plugins.BF')
        options = sj.jimport('loci.plugins.in.ImporterOptions')() # import and initialize ImporterOptions
        options.setId(image_url)
        options.setVirtual(False)
        options.setSplitChannels(True)

        options.setCBegin(0,1)
        options.setCEnd(0,1)
        options.setCStep(0,1)

        options.setZBegin(0,3)
        options.setZEnd(0,12)
        options.setZStep(0,1)

        options.setTBegin(0,0)
        options.setTEnd(0,0)
        options.setTStep(0,1)
        imp = BF.openImagePlus(options)[0]
        imp.show();
        self.ij.IJ.run("Stack to Images", "");
        counter = 0
        while(True):
            counter = counter + 1
            activeImage = self.ij.py.active_imageplus()
            if (activeImage == None):
                break
            else:
                print(activeImage.getFileInfo())
                self.ij.IJ.saveAs(activeImage, "Tiff", "A:/Development/NaoualImageAnalyzer/data/outputs/Experiment-1022_" + str(counter) +".tif");
                activeImage.close()